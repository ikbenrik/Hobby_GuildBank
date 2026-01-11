"""
bot/services/backup_service.py

Service responsible for exporting and restoring Google Sheets data
used by the Guild Bank bot.

Responsibilities:
- Export the entire spreadsheet as an XLSX backup
- Restore selected worksheets from an XLSX file back into Google Sheets

This service contains NO Discord-specific logic.
It is intended to be used by backup/restore cogs only.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any
from datetime import datetime
import requests

from google.auth.transport.requests import Request
from google.oauth2.service_account import Credentials as ServiceAccountCredentials
import openpyxl


@dataclass
class RestoreTarget:
    """
    Describes a single worksheet restore target.

    sheet_name:
        Name of the worksheet inside the XLSX file.
    worksheet:
        gspread worksheet object to overwrite.
    """
    sheet_name: str
    worksheet: Any


class BackupService:
    """
    Handles spreadsheet-level backup and restore operations.

    This class:
    - Authenticates using a Google service account
    - Downloads the spreadsheet as XLSX
    - Restores specific sheets back into Google Sheets
    """

    def __init__(self, spreadsheet_id: str, creds_file: str):
        # Store spreadsheet ID and credentials file path
        self.spreadsheet_id = spreadsheet_id
        self.creds_file = creds_file

    def export_xlsx(self) -> str:
        """
        Export the Google Spreadsheet to a timestamped XLSX file.

        Returns:
            str: Filename of the generated XLSX backup.
        """
        now = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"GuildBankBackup_{now}.xlsx"

        # Google Sheets export endpoint
        export_url = (
            f"https://docs.google.com/spreadsheets/"
            f"d/{self.spreadsheet_id}/export?format=xlsx"
        )

        # Required scope for spreadsheet access
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]

        # Load service account credentials
        creds = ServiceAccountCredentials.from_service_account_file(
            self.creds_file,
            scopes=scopes
        )

        # Ensure access token is valid
        if not creds.token:
            creds.refresh(Request())

        # Authorized request to Google export endpoint
        headers = {"Authorization": f"Bearer {creds.token}"}
        resp = requests.get(export_url, headers=headers)
        resp.raise_for_status()

        # Write XLSX file to disk
        with open(filename, "wb") as f:
            f.write(resp.content)

        return filename

    def restore_xlsx_to_sheets(self, filename: str, targets: list[RestoreTarget]) -> None:
        """
        Restore selected worksheets from an XLSX file back into Google Sheets.

        Args:
            filename:
                Path to the XLSX backup file.
            targets:
                List of RestoreTarget objects defining which sheets to restore.
        """
        # Load XLSX workbook from disk
        wb = openpyxl.load_workbook(filename)

        for t in targets:
            # Skip sheets that do not exist in the backup
            if t.sheet_name not in wb.sheetnames:
                continue

            sheet = wb[t.sheet_name]

            # Convert worksheet into a 2D list of values
            data = [[cell.value for cell in row] for row in sheet.iter_rows()]
            if not data:
                continue

            # Clear destination worksheet and overwrite with backup data
            t.worksheet.clear()
            t.worksheet.update(range_name="A1", values=data)
